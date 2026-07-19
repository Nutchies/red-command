from fastapi import APIRouter, Depends, HTTPException, status, Header, UploadFile, File, Form
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import Response, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import func
from sqlalchemy import or_
from jose import JWTError, jwt
from datetime import datetime, timedelta
from typing import List, Optional
from pydantic import BaseModel
import re
import bcrypt
from cryptography.fernet import Fernet
import base64
import io
import random
import string
from PIL import Image, ImageDraw, ImageFont

from app.db.database import get_db
from app.core.config import settings

ENCRYPTION_KEY = base64.urlsafe_b64encode(settings.SECRET_KEY[:32].encode('utf-8'))
fernet = Fernet(ENCRYPTION_KEY)
from app.models.schemas import (
    Token, HeartbeatRequest, SyncRequest, ClientResponse,
    ActionResponse, AIExtractedResponse, DashboardStats, UserCreate, UserResponse,
    UserCreateRequest, UserUpdateRequest,
    VideoResponse, PenTestResultResponse, ToolResponse, ToolVersionResponse,
    TaskPlanResponse, TaskTargetResponse,
    AssetResponse, AssetCreateRequest, AssetUpdateRequest,
    ChatRoomResponse, ChatRoomCreateRequest, ChatMessageResponse, ChatMessageCreateRequest, AddChatMemberRequest
)
from app.models.models import User, Client, Action, AIExtracted, Video, PenTestResult, Tool, ToolVersion, TaskPlan, TaskTarget, Asset, ChatRoom, ChatMessage, ChatRoomMember
from app.services.services import ClientService, ActionService

router = APIRouter()
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/token")

captcha_store = {}


def generate_captcha_text(length: int = 4) -> str:
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))


def generate_captcha_image(text: str) -> bytes:
    width, height = 120, 40
    image = Image.new('RGB', (width, height), (243, 244, 246))
    draw = ImageDraw.Draw(image)
    
    try:
        font = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 28)
    except:
        font = ImageFont.load_default()
    
    chars = list(text)
    for i, char in enumerate(chars):
        x = 15 + i * 25 + random.randint(-3, 3)
        y = 5 + random.randint(-3, 3)
        color = (random.randint(50, 150), random.randint(50, 150), random.randint(50, 150))
        draw.text((x, y), char, font=font, fill=color)
    
    for _ in range(20):
        x1, y1 = random.randint(0, width), random.randint(0, height)
        x2, y2 = random.randint(0, width), random.randint(0, height)
        draw.line((x1, y1, x2, y2), fill=(random.randint(100, 200), random.randint(100, 200), random.randint(100, 200)), width=1)
    
    for _ in range(50):
        x, y = random.randint(0, width), random.randint(0, height)
        draw.point((x, y), fill=(random.randint(50, 200), random.randint(50, 200), random.randint(50, 200)))
    
    buf = io.BytesIO()
    image.save(buf, format='PNG')
    return buf.getvalue()


@router.get("/captcha")
async def get_captcha():
    captcha_id = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
    captcha_text = generate_captcha_text()
    captcha_store[captcha_id] = {"text": captcha_text.lower(), "expire": datetime.utcnow() + timedelta(minutes=5)}
    
    image_data = generate_captcha_image(captcha_text)
    image_base64 = base64.b64encode(image_data).decode('utf-8')
    
    for key in list(captcha_store.keys()):
        if captcha_store[key]["expire"] < datetime.utcnow():
            del captcha_store[key]
    
    return {"captcha_id": captcha_id, "image_base64": image_base64}


def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.SECRET_KEY, algorithm=settings.ALGORITHM)


async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise credentials_exception
    return user


async def get_current_admin(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized: Admin access required"
        )
    return current_user


def verify_password(stored_password: str, provided_password: str) -> bool:
    if stored_password.startswith('$2b$'):
        return bcrypt.checkpw(provided_password.encode('utf-8'), stored_password.encode('utf-8'))
    else:
        try:
            decrypted = fernet.decrypt(stored_password.encode('utf-8')).decode('utf-8')
            return decrypted == provided_password
        except:
            return stored_password == provided_password


@router.post("/token", response_model=Token)
async def login(
    username: str = Form(...),
    password: str = Form(...),
    captcha_id: str = Form(...),
    captcha_code: str = Form(...),
    db: Session = Depends(get_db)
):
    captcha_data = captcha_store.get(captcha_id)
    if not captcha_data:
        raise HTTPException(status_code=400, detail="验证码已过期，请刷新验证码")
    
    if captcha_data["expire"] < datetime.utcnow():
        del captcha_store[captcha_id]
        raise HTTPException(status_code=400, detail="验证码已过期，请刷新验证码")
    
    if captcha_code.lower() != captcha_data["text"]:
        raise HTTPException(status_code=400, detail="验证码错误")
    
    del captcha_store[captcha_id]
    
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=400, detail="用户名或密码错误")
    
    if not verify_password(user.password_hash, password):
        raise HTTPException(status_code=400, detail="用户名或密码错误")

    access_token = create_access_token(data={"sub": user.username})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "username": user.username,
        "role": user.role,
        "user_group": user.user_group,
        "user_id": user.id
    }


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@router.post("/change-password")
async def change_password(
    request: ChangePasswordRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not verify_password(current_user.password_hash, request.old_password):
        raise HTTPException(status_code=400, detail="原密码错误")
    
    if len(request.new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码长度不能少于6位")
    
    if request.old_password == request.new_password:
        raise HTTPException(status_code=400, detail="新密码不能与原密码相同")
    
    encrypted_password = encrypt_password(request.new_password)
    current_user.password_hash = encrypted_password
    db.commit()
    
    return {"message": "密码修改成功"}


def encrypt_password(password: str) -> str:
    return fernet.encrypt(password.encode('utf-8')).decode('utf-8')


def decrypt_password(stored_password: str) -> Optional[str]:
    if stored_password.startswith('$2b$'):
        return None
    try:
        return fernet.decrypt(stored_password.encode('utf-8')).decode('utf-8')
    except:
        return None


@router.post("/register", response_model=UserResponse)
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == user_data.username).first():
        raise HTTPException(status_code=400, detail="Username already exists")

    encrypted_password = encrypt_password(user_data.password)
    user = User(username=user_data.username, password_hash=encrypted_password)
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.get("/users", response_model=List[UserResponse])
async def get_users(db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    users = db.query(User).all()
    result = []
    for user in users:
        user_response = UserResponse.from_orm(user)
        user_response.password = decrypt_password(user.password_hash)
        result.append(user_response)
    return result


@router.get("/users/list", response_model=List[dict])
async def get_user_list(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    users = db.query(User.id, User.username, User.user_group).all()
    return [{"id": user.id, "username": user.username, "user_group": user.user_group} for user in users]


@router.post("/users", response_model=UserResponse)
async def create_user(user_data: UserCreateRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    if db.query(User).filter(User.username == user_data.username).first():
        raise HTTPException(status_code=400, detail="Username already exists")

    encrypted_password = encrypt_password(user_data.password)
    user = User(
        username=user_data.username,
        password_hash=encrypted_password,
        role=user_data.role or "operator",
        user_group=user_data.user_group or "未分组",
        organization=user_data.organization or ""
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    user_response = UserResponse.from_orm(user)
    user_response.password = user_data.password
    return user_response


@router.get("/users/groups")
async def get_user_groups(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    groups = db.query(User.user_group).distinct().all()
    return [g[0] for g in groups]


@router.get("/users/{user_id}", response_model=UserResponse)
async def get_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


@router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: int, update_data: UserUpdateRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if update_data.role is not None:
        user.role = update_data.role
    if update_data.user_group is not None:
        user.user_group = update_data.user_group
    if update_data.is_active is not None:
        user.is_active = update_data.is_active

    db.commit()
    db.refresh(user)
    return user


@router.delete("/users/{user_id}")
async def delete_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    db.delete(user)
    db.commit()
    return {"status": "ok", "message": "用户删除成功"}


@router.post("/heartbeat")
async def heartbeat(data: HeartbeatRequest, db: Session = Depends(get_db)):
    client_service = ClientService(db)
    client = client_service.update_heartbeat(data.client_id, data.hostname, data.ip, data.version)
    return {"status": "ok", "client_id": client.client_id}


@router.post("/actions")
async def sync_actions(data: SyncRequest, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.client_id == data.client_id).first()
    if not client:
        client_service = ClientService(db)
        client = client_service.get_or_create_client(data.client_id, "unknown", "127.0.0.1", "1.0.0")

    action_service = ActionService(db)
    actions = action_service.create_actions_batch(client.id, [a.dict() for a in data.actions])

    return {"status": "ok", "received": len(actions)}


@router.get("/clients", response_model=List[ClientResponse])
async def get_clients(db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    client_service = ClientService(db)
    client_service.check_offline_clients()
    return client_service.get_all_clients()


@router.get("/clients/search", response_model=List[ClientResponse])
async def search_clients(
    keyword: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    from app.models.models import Client
    from sqlalchemy import or_
    client_service = ClientService(db)
    client_service.check_offline_clients()
    
    if not keyword.strip():
        return client_service.get_all_clients()
    
    clients = db.query(Client).filter(
        or_(
            Client.ip_address.like(f"%{keyword}%"),
            Client.client_id.like(f"%{keyword}%"),
            Client.hostname.like(f"%{keyword}%"),
        )
    ).order_by(Client.last_heartbeat.desc()).all()
    
    return clients


@router.delete("/clients/{client_id}")
def delete_client(client_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    from sqlalchemy import text
    try:
        int_id = int(client_id)
        result = db.execute(text('SELECT id FROM clients WHERE id = :client_id'), {'client_id': int_id})
        row = result.fetchone()
        if row:
            db.execute(text('DELETE FROM actions WHERE client_id = :client_id'), {'client_id': int_id})
            db.execute(text('DELETE FROM clients WHERE id = :client_id'), {'client_id': int_id})
            db.commit()
            return {"status": "ok", "message": "Client deleted"}
        else:
            result2 = db.execute(text('SELECT id FROM clients WHERE client_id = :client_id'), {'client_id': client_id})
            row2 = result2.fetchone()
            if row2:
                db.execute(text('DELETE FROM actions WHERE client_id = :client_id'), {'client_id': row2[0]})
                db.execute(text('DELETE FROM clients WHERE id = :client_id'), {'client_id': row2[0]})
                db.commit()
                return {"status": "ok", "message": "Client deleted"}
            raise HTTPException(status_code=404, detail="Client not found")
    except ValueError:
        result = db.execute(text('SELECT id FROM clients WHERE client_id = :client_id'), {'client_id': client_id})
        row = result.fetchone()
        if row:
            db.execute(text('DELETE FROM actions WHERE client_id = :client_id'), {'client_id': row[0]})
            db.execute(text('DELETE FROM clients WHERE id = :client_id'), {'client_id': row[0]})
            db.commit()
            return {"status": "ok", "message": "Client deleted"}
        raise HTTPException(status_code=404, detail="Client not found")


class ClientUpdateRequest(BaseModel):
    organization: Optional[str] = None
    hostname: Optional[str] = None


@router.put("/clients/{client_id}", response_model=ClientResponse)
async def update_client(
    client_id: str,
    data: ClientUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    from sqlalchemy import text
    try:
        int_id = int(client_id)
        result = db.execute(text('SELECT id FROM clients WHERE id = :client_id'), {'client_id': int_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Client not found")
        client_id_int = int_id
    except ValueError:
        result = db.execute(text('SELECT id FROM clients WHERE client_id = :client_id'), {'client_id': client_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Client not found")
        client_id_int = row[0]

    update_data = {}
    if data.organization is not None:
        update_data['organization'] = data.organization
    if data.hostname is not None:
        update_data['hostname'] = data.hostname

    if update_data:
        set_clause = ', '.join([f"{k} = :{k}" for k in update_data.keys()])
        update_data['client_id'] = client_id_int
        db.execute(text(f'UPDATE clients SET {set_clause} WHERE id = :client_id'), update_data)
        db.commit()

    client = db.query(Client).filter(Client.id == client_id_int).first()
    return ClientResponse.from_orm(client)


@router.get("/clients/{client_id}/actions", response_model=List[ActionResponse])
async def get_client_actions(
    client_id: str,
    limit: int = 100,
    action_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    client = db.query(Client).filter(Client.client_id == client_id).first()
    if not client:
        return []
    query = db.query(Action).filter(Action.client_id == client.id)
    if action_type:
        query = query.filter(Action.action_type == action_type)
    actions = query.order_by(Action.timestamp.desc()).limit(limit).all()
    
    result = []
    for action in actions:
        action_dict = {
            "id": action.id,
            "client_id": action.client_id,
            "action_type": action.action_type,
            "content": action.content,
            "result": action.result,
            "exit_code": action.exit_code,
            "timestamp": action.timestamp,
            "created_at": action.created_at,
            "client_hostname": action.client.hostname if action.client else None
        }
        result.append(ActionResponse(**action_dict))
    return result


@router.get("/actions", response_model=List[ActionResponse])
async def get_actions(
    limit: int = 100,
    offset: int = 0,
    action_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    action_service = ActionService(db)
    actions = action_service.get_actions(limit, offset, action_type=action_type)

    result = []
    for action in actions:
        action_dict = {
            "id": action.id,
            "client_id": action.client_id,
            "action_type": action.action_type,
            "content": action.content,
            "result": action.result,
            "exit_code": action.exit_code,
            "timestamp": action.timestamp,
            "created_at": action.created_at,
            "client_hostname": action.client.hostname if action.client else None
        }
        result.append(ActionResponse(**action_dict))
    return result


@router.get("/ai/extracted", response_model=List[AIExtractedResponse])
async def get_extracted_info(
    info_type: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(AIExtracted)
    if info_type:
        query = query.filter(AIExtracted.type == info_type)
    return query.order_by(AIExtracted.extracted_at.desc()).limit(limit).all()


@router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_admin)):
    client_service = ClientService(db)
    action_service = ActionService(db)

    client_service.check_offline_clients()

    from app.services.ai_analyzer import AIAnalyzer
    ai_analyzer = AIAnalyzer(db)
    ai_stats = ai_analyzer.get_statistics()

    total_clients = db.query(Client).count()
    online_clients = db.query(Client).filter(Client.status == "online").count()

    from app.models.models import PenTestResult, Tool
    return DashboardStats(
        total_clients=total_clients,
        online_clients=online_clients,
        offline_clients=total_clients - online_clients,
        total_actions=db.query(Action).count(),
        actions_today=action_service.get_actions_today_count(),
        pen_test_count=db.query(PenTestResult).count(),
        tool_count=db.query(Tool).count(),
        ports_found=ai_stats.get("ports", 0),
        services_found=ai_stats.get("ports", 0),
        vulnerabilities_found=ai_stats.get("vulnerabilities", 0),
        credentials_found=ai_stats.get("credentials", 0)
    )


import os
import base64
from fastapi import File, UploadFile, Form
from fastapi.responses import FileResponse


VIDEO_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "videos")
os.makedirs(VIDEO_DIR, exist_ok=True)


@router.post("/videos/upload")
async def upload_video(
    client_id: str = Form(""),
    session_id: str = Form(""),
    nonce: str = Form(""),
    duration: float = Form(0),
    video: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not client_id or not session_id:
        raise HTTPException(status_code=400, detail="缺少必要参数")

    client = db.query(Client).filter(Client.client_id == client_id).first()
    if not client:
        client_service = ClientService(db)
        client = client_service.get_or_create_client(client_id, "unknown", "127.0.0.1", "1.0.0")

    file_path = os.path.join(VIDEO_DIR, f"{session_id}.mp4.enc")
    
    with open(file_path, "wb") as f:
        content = await video.read()
        f.write(content)
    
    video_record = Video(
        client_id=client.id,
        session_id=session_id,
        file_path=file_path,
        nonce=nonce,
        file_size=len(content),
        duration=duration,
        timestamp=datetime.now().timestamp()
    )
    db.add(video_record)
    db.commit()
    db.refresh(video_record)

    return {"status": "ok", "id": video_record.id, "duration": duration}


@router.get("/clients/{client_id}/videos", response_model=List[VideoResponse])
async def get_client_videos(
    client_id: str,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    client = db.query(Client).filter(Client.client_id == client_id).first()
    if not client:
        return []
    
    videos = db.query(Video).filter(Video.client_id == client.id).order_by(Video.timestamp.desc()).limit(limit).all()
    
    results = []
    for video in videos:
        response = VideoResponse.from_orm(video)
        response.client_id = client.client_id
        response.ip_address = client.ip_address
        response.hostname = client.hostname
        results.append(response)
    
    return results


@router.get("/videos/search", response_model=List[VideoResponse])
async def search_videos(
    keyword: Optional[str] = "",
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    results = []
    
    if keyword:
        keyword = keyword.strip()
        if len(keyword) < 3:
            clients = db.query(Client).filter(Client.client_id == keyword).all()
        else:
            clients = db.query(Client).filter(
                (Client.client_id == keyword) |
                (Client.hostname.like(f"%{keyword}%")) |
                (Client.ip_address.like(f"%{keyword}%"))
            ).all()
        
        if not clients and len(keyword) >= 3:
            videos_by_session = db.query(Video).filter(
                Video.session_id.like(f"%{keyword}%")
            ).all()
            
            client_ids_from_videos = set()
            for video in videos_by_session:
                client_ids_from_videos.add(video.client_id)
            
            for cid in client_ids_from_videos:
                client = db.query(Client).filter(Client.id == cid).first()
                if client:
                    clients.append(client)
    else:
        clients = db.query(Client).all()
    
    for client in clients:
        videos = db.query(Video).filter(Video.client_id == client.id).order_by(Video.timestamp.desc()).limit(limit).all()
        
        for video in videos:
            response = VideoResponse.from_orm(video)
            response.client_id = client.client_id
            response.ip_address = client.ip_address
            response.hostname = client.hostname
            results.append(response)
    
    results.sort(key=lambda x: x.timestamp, reverse=True)
    return results[:limit]


@router.get("/videos/{video_id}")
async def download_video(
    video_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    video = db.query(Video).filter(Video.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="视频不存在")
    
    if not os.path.exists(video.file_path):
        raise HTTPException(status_code=404, detail="视频文件已删除")
    
    return FileResponse(
        video.file_path,
        media_type="application/octet-stream",
        filename=os.path.basename(video.file_path)
    )


@router.get("/videos/{video_id}/play")
async def play_video(
    video_id: int,
    token: str = "",
    db: Session = Depends(get_db),
    range: str = Header(None)
):
    import base64
    from jose import JWTError, jwt
    from fastapi.responses import StreamingResponse
    import io
    
    if token:
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            username: str = payload.get("sub")
            if username is None:
                raise HTTPException(status_code=401, detail="Invalid token")
        except JWTError:
            raise HTTPException(status_code=401, detail="Invalid token")
    else:
        raise HTTPException(status_code=401, detail="Missing token")
    
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    
    video = db.query(Video).filter(Video.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="视频不存在")
    
    if not os.path.exists(video.file_path):
        raise HTTPException(status_code=404, detail="视频文件已删除")
    
    encryption_key = b'redteam2024!@#$%^&*()_+-=[]{}|;\':",./<>?'[:32]
    
    with open(video.file_path, "rb") as f:
        encrypted_data = f.read()
    
    nonce = base64.b64decode(video.nonce)
    
    try:
        aesgcm = AESGCM(encryption_key)
        decrypted_data = aesgcm.decrypt(nonce, encrypted_data, None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解密失败: {str(e)}")
    
    total_size = len(decrypted_data)
    content_type = "video/mp4"
    
    if range:
        range_header = range.strip()
        range_match = re.match(r'bytes=(\d+)-(\d*)', range_header)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2)) if range_match.group(2) else total_size - 1
            
            if start >= total_size:
                return Response(status_code=416)
            
            chunk_size = end - start + 1
            chunk = decrypted_data[start:end+1]
            
            return Response(
                content=chunk,
                status_code=206,
                media_type=content_type,
                headers={
                    "Content-Range": f"bytes {start}-{end}/{total_size}",
                    "Content-Length": str(chunk_size),
                    "Accept-Ranges": "bytes",
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, OPTIONS",
                    "Access-Control-Allow-Headers": "Range"
                }
            )
    
    return Response(
        content=decrypted_data,
        media_type=content_type,
        headers={
            "Content-Length": str(total_size),
            "Accept-Ranges": "bytes",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "Range"
        }
    )


class BatchDeleteRequest(BaseModel):
    ids: List[int]


@router.delete("/videos/batch")
async def batch_delete_videos(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not request.ids:
        raise HTTPException(status_code=400, detail="请选择要删除的视频")
    
    deleted_count = 0
    for video_id in request.ids:
        video = db.query(Video).filter(Video.id == video_id).first()
        if video:
            file_path = video.file_path
            if os.path.exists(file_path):
                os.remove(file_path)
            db.delete(video)
            deleted_count += 1
    
    db.commit()
    
    return {"status": "ok", "message": f"成功删除 {deleted_count} 个视频", "deleted_count": deleted_count}


@router.delete("/videos/{video_id}")
async def delete_video(
    video_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    video = db.query(Video).filter(Video.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="视频不存在")
    
    file_path = video.file_path
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.delete(video)
    db.commit()
    
    return {"status": "ok", "message": "视频删除成功"}


UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.post("/pen-test/upload")
async def upload_pen_test_result(
    target_ip: str = Form(...),
    target_organization: Optional[str] = Form(None),
    attacker_ip: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    category: Optional[str] = Form("other"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import uuid
    
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    try:
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")
    
    content_type = file.content_type
    if content_type:
        file_type = content_type.split("/")[0]
    else:
        file_type = "other"
    
    result = PenTestResult(
        target_ip=target_ip,
        target_organization=target_organization,
        attacker_ip=attacker_ip,
        file_name=file.filename,
        file_path=file_path,
        file_size=len(contents),
        file_type=file_type,
        remark=remark,
        category=category,
        user_group=current_user.user_group,
        created_by=current_user.id
    )
    
    db.add(result)
    db.commit()
    db.refresh(result)
    
    return PenTestResultResponse.from_orm(result)


@router.get("/pen-test/search", response_model=List[PenTestResultResponse])
async def search_pen_test_results(
    keyword: Optional[str] = "",
    target_ip: Optional[str] = "",
    category: Optional[str] = "",
    user_group: Optional[str] = "",
    created_by_username: Optional[str] = "",
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(PenTestResult, User.username.label("created_by_username"), User.organization.label("created_by_organization"))\
              .outerjoin(User, PenTestResult.created_by == User.id)
    
    if current_user.role != "admin":
        query = query.filter(PenTestResult.created_by == current_user.id)
    
    if keyword:
        query = query.filter(or_(
            PenTestResult.file_name.like(f"%{keyword}%"),
            PenTestResult.target_ip.like(f"%{keyword}%"),
            User.username.like(f"%{keyword}%"),
            PenTestResult.user_group.like(f"%{keyword}%")
        ))
    else:
        if target_ip:
            query = query.filter(PenTestResult.target_ip.like(f"%{target_ip}%"))
        
        if category:
            query = query.filter(PenTestResult.category == category)
        
        if user_group:
            query = query.filter(PenTestResult.user_group.like(f"%{user_group}%"))
        
        if created_by_username:
            query = query.filter(User.username.like(f"%{created_by_username}%"))
    
    results = query.order_by(PenTestResult.created_at.desc()).limit(limit).all()
    
    response_list = []
    for result, username, organization in results:
        response = PenTestResultResponse.from_orm(result)
        response.created_by_username = username
        response.created_by_organization = organization
        response_list.append(response)
    
    return response_list


@router.get("/pen-test/{result_id}")
async def download_pen_test_result(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = db.query(PenTestResult).filter(PenTestResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="成果不存在")
    
    if not os.path.exists(result.file_path):
        raise HTTPException(status_code=404, detail="文件已删除")
    
    return FileResponse(
        result.file_path,
        media_type="application/octet-stream",
        filename=result.file_name
    )


@router.delete("/pen-test/{result_id}")
async def delete_pen_test_result(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = db.query(PenTestResult).filter(PenTestResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="成果不存在")
    
    if current_user.role != "admin" and result.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除该成果")
    
    file_path = result.file_path
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.delete(result)
    db.commit()
    
    return {"status": "ok", "message": "成果删除成功"}


@router.delete("/pen-test/batch/{target_ip}")
async def delete_pen_test_host(
    target_ip: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(PenTestResult).filter(PenTestResult.target_ip == target_ip)
    
    if current_user.role != "admin":
        query = query.filter(PenTestResult.created_by == current_user.id)
    
    results = query.all()
    if not results:
        raise HTTPException(status_code=404, detail="未找到该IP的成果记录")
    
    for result in results:
        if result.file_path and os.path.exists(result.file_path):
            os.remove(result.file_path)
        db.delete(result)
    
    db.commit()
    
    return {"status": "ok", "message": f"成功删除 {len(results)} 条成果记录"}


TOOLS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "tools")
os.makedirs(TOOLS_DIR, exist_ok=True)


@router.post("/tools")
async def create_tool(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    category: Optional[str] = Form("other"),
    url: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    existing = db.query(Tool).filter(Tool.name == name).first()
    if existing:
        raise HTTPException(status_code=400, detail="工具名称已存在")
    
    tool = Tool(
        name=name,
        description=description,
        category=category,
        url=url
    )
    db.add(tool)
    db.commit()
    db.refresh(tool)
    
    return ToolResponse.from_orm(tool)


@router.get("/tools", response_model=List[ToolResponse])
async def get_tools(
    keyword: Optional[str] = "",
    category: Optional[str] = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Tool)
    
    if keyword:
        query = query.filter(Tool.name.like(f"%{keyword}%"))
    
    if category:
        query = query.filter(Tool.category == category)
    
    tools = query.order_by(Tool.name).all()
    return [ToolResponse.from_orm(t) for t in tools]


@router.get("/tools/{tool_id}", response_model=ToolResponse)
async def get_tool(
    tool_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    return ToolResponse.from_orm(tool)


@router.put("/tools/{tool_id}", response_model=ToolResponse)
async def update_tool(
    tool_id: int,
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    url: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    
    if name:
        existing = db.query(Tool).filter(Tool.name == name, Tool.id != tool_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="工具名称已存在")
        tool.name = name
    
    if description is not None:
        tool.description = description
    
    if category is not None:
        tool.category = category
    
    if url is not None:
        tool.url = url
    
    db.commit()
    db.refresh(tool)
    
    return ToolResponse.from_orm(tool)


@router.delete("/tools/{tool_id}")
async def delete_tool(
    tool_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    
    for version in tool.versions:
        if version.file_path and os.path.exists(version.file_path):
            os.remove(version.file_path)
    
    db.delete(tool)
    db.commit()
    
    return {"status": "ok", "message": "工具删除成功"}


@router.post("/tools/{tool_id}/versions")
async def upload_tool_version(
    tool_id: int,
    version: str = Form(...),
    platform: Optional[str] = Form("linux"),
    url: Optional[str] = Form(None),
    username: Optional[str] = Form(None),
    password: Optional[str] = Form(None),
    user_id: Optional[int] = Form(None),
    user_group: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    
    existing = db.query(ToolVersion).filter(
        ToolVersion.tool_id == tool_id,
        ToolVersion.version == version,
        ToolVersion.platform == platform
    ).first()
    if existing:
        if existing.file_path and os.path.exists(existing.file_path):
            os.remove(existing.file_path)
    
    file_path = None
    file_name = None
    file_size = 0
    
    if file:
        import uuid
        file_ext = os.path.splitext(file.filename)[1]
        unique_filename = f"{tool.name}_{version}_{platform}{file_ext}"
        file_path = os.path.join(TOOLS_DIR, unique_filename)
        
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        
        file_name = file.filename
        file_size = len(contents)
    
    if user_id is None:
        user_id = current_user.id
    if user_group is None:
        user_group = current_user.user_group
    
    tool_version = ToolVersion(
        tool_id=tool_id,
        version=version,
        file_path=file_path,
        file_name=file_name,
        file_size=file_size,
        platform=platform,
        url=url,
        username=username,
        password=password,
        user_id=user_id,
        user_group=user_group
    )
    
    db.add(tool_version)
    db.commit()
    db.refresh(tool_version)
    
    return ToolVersionResponse.from_orm(tool_version)


@router.get("/tools/{tool_id}/versions", response_model=List[ToolVersionResponse])
async def get_tool_versions(
    tool_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    
    query = db.query(ToolVersion).filter(ToolVersion.tool_id == tool_id)
    
    if current_user.role != "admin":
        query = query.filter(
            or_(
                ToolVersion.user_id == current_user.id,
                ToolVersion.user_group == current_user.user_group
            )
        )
    
    versions = query.order_by(ToolVersion.version.desc()).all()
    
    return [ToolVersionResponse.from_orm(v) for v in versions]


@router.get("/tools/download/{version_id}")
async def download_tool_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    version = db.query(ToolVersion).filter(ToolVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")
    
    if not version.file_path or not os.path.exists(version.file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return FileResponse(
        version.file_path,
        media_type="application/octet-stream",
        filename=version.file_name or os.path.basename(version.file_path)
    )


@router.delete("/tools/versions/{version_id}")
async def delete_tool_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    version = db.query(ToolVersion).filter(ToolVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")
    
    if version.file_path and os.path.exists(version.file_path):
        os.remove(version.file_path)
    
    db.delete(version)
    db.commit()
    
    return {"status": "ok", "message": "版本删除成功"}


@router.post("/task-plans", response_model=TaskPlanResponse)
async def create_task_plan(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    team: str = Form(...),
    status: Optional[str] = Form("pending"),
    start_time: Optional[str] = Form(None),
    end_time: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = TaskPlan(
        name=name,
        description=description,
        team=team,
        status=status,
        start_time=datetime.fromisoformat(start_time) if start_time else None,
        end_time=datetime.fromisoformat(end_time) if end_time else None,
        created_by=current_user.id
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    
    return TaskPlanResponse.from_orm(plan)


@router.get("/task-plans", response_model=List[TaskPlanResponse])
async def get_task_plans(
    keyword: Optional[str] = "",
    status: Optional[str] = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(TaskPlan)
    
    if current_user.role != "admin":
        query = query.filter(
            or_(
                TaskPlan.created_by == current_user.id,
                TaskPlan.team == current_user.user_group
            )
        )
    
    if keyword:
        query = query.filter(
            TaskPlan.id.in_(
                db.query(TaskTarget.plan_id).filter(
                    or_(
                        TaskTarget.target_value.like(f"%{keyword}%"),
                        TaskTarget.assigned_team.like(f"%{keyword}%")
                    )
                )
            )
        )
    
    if status:
        query = query.filter(TaskPlan.status == status)
    
    plans = query.order_by(TaskPlan.created_at.desc()).all()
    return [TaskPlanResponse.from_orm(p) for p in plans]


@router.get("/task-plans/active", response_model=List[TaskPlanResponse])
async def get_active_task_plans(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(TaskPlan).filter(
        TaskPlan.status.in_(["pending", "active", "running"])
    ).order_by(TaskPlan.created_at.desc())
    
    if current_user.role != "admin":
        query = query.filter(
            or_(
                TaskPlan.created_by == current_user.id,
                TaskPlan.team == current_user.user_group
            )
        )
    
    return [TaskPlanResponse.from_orm(p) for p in query.all()]


@router.get("/task-plans/{plan_id}", response_model=TaskPlanResponse)
async def get_task_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = db.query(TaskPlan).filter(TaskPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="任务管理不存在")
    return TaskPlanResponse.from_orm(plan)


@router.put("/task-plans/{plan_id}", response_model=TaskPlanResponse)
async def update_task_plan(
    plan_id: int,
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    team: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    start_time: Optional[str] = Form(None),
    end_time: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = db.query(TaskPlan).filter(TaskPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="任务管理不存在")
    
    if name:
        plan.name = name
    if description is not None:
        plan.description = description
    if team:
        plan.team = team
    if status:
        plan.status = status
    if start_time:
        plan.start_time = datetime.fromisoformat(start_time)
    if end_time:
        plan.end_time = datetime.fromisoformat(end_time)
    
    db.commit()
    db.refresh(plan)
    
    return TaskPlanResponse.from_orm(plan)


@router.delete("/task-plans/{plan_id}")
async def delete_task_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = db.query(TaskPlan).filter(TaskPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="任务管理不存在")
    
    if current_user.role != "admin" and plan.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除该任务")
    
    db.delete(plan)
    db.commit()
    
    return {"status": "ok", "message": "任务管理删除成功"}


@router.post("/task-plans/{plan_id}/targets", response_model=TaskTargetResponse)
async def add_task_target(
    plan_id: int,
    target_value: str = Form(...),
    target_organization: Optional[str] = Form(None),
    progress: Optional[int] = Form(0),
    organization: Optional[str] = Form(None),
    assigned_team: Optional[str] = Form(None),
    start_time: Optional[str] = Form(None),
    end_time: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = db.query(TaskPlan).filter(TaskPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="任务管理不存在")
    
    from datetime import datetime as dt
    
    start_time_dt = None
    end_time_dt = None
    if start_time:
        try:
            start_time_dt = dt.strptime(start_time, "%Y-%m-%d %H:%M:%S")
        except:
            pass
    if end_time:
        try:
            end_time_dt = dt.strptime(end_time, "%Y-%m-%d %H:%M:%S")
        except:
            pass
    
    target = TaskTarget(
        plan_id=plan_id,
        target_value=target_value,
        target_organization=target_organization,
        progress=progress,
        organization=organization,
        assigned_team=assigned_team,
        start_time=start_time_dt,
        end_time=end_time_dt
    )
    db.add(target)
    db.commit()
    db.refresh(target)
    return TaskTargetResponse.from_orm(target)


@router.get("/task-plans/{plan_id}/targets", response_model=List[TaskTargetResponse])
async def get_task_targets(
    plan_id: int,
    keyword: Optional[str] = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    plan = db.query(TaskPlan).filter(TaskPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="任务管理不存在")
    
    query = db.query(TaskTarget).filter(TaskTarget.plan_id == plan_id)
    
    if keyword:
        query = query.filter(or_(
            TaskTarget.target_value.like(f"%{keyword}%"),
            TaskTarget.assigned_team.like(f"%{keyword}%")
        ))
    
    targets = query.order_by(TaskTarget.created_at.desc()).all()
    
    return [TaskTargetResponse.from_orm(t) for t in targets]


@router.put("/task-targets/{target_id}", response_model=TaskTargetResponse)
async def update_task_target(
    target_id: int,
    target_value: Optional[str] = Form(None),
    target_organization: Optional[str] = Form(None),
    progress: Optional[int] = Form(None),
    organization: Optional[str] = Form(None),
    assigned_team: Optional[str] = Form(None),
    start_time: Optional[str] = Form(None),
    end_time: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    target = db.query(TaskTarget).filter(TaskTarget.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="目标不存在")
    
    from datetime import datetime as dt
    
    if target_value:
        target.target_value = target_value
    if target_organization:
        target.target_organization = target_organization
    if progress is not None:
        target.progress = progress
    if organization:
        target.organization = organization
    if assigned_team:
        target.assigned_team = assigned_team
    if start_time:
        try:
            target.start_time = dt.strptime(start_time, "%Y-%m-%d %H:%M:%S")
        except:
            pass
    if end_time:
        try:
            target.end_time = dt.strptime(end_time, "%Y-%m-%d %H:%M:%S")
        except:
            pass
    
    db.commit()
    db.refresh(target)
    return TaskTargetResponse.from_orm(target)


@router.delete("/task-targets/{target_id}")
async def delete_task_target(
    target_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    target = db.query(TaskTarget).filter(TaskTarget.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="目标不存在")
    
    plan = db.query(TaskPlan).filter(TaskPlan.id == target.plan_id).first()
    if current_user.role != "admin" and plan.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除该目标")
    
    db.delete(target)
    db.commit()
    
    return {"status": "ok", "message": "目标删除成功"}


@router.get("/teams")
async def get_teams(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    teams = db.query(TaskPlan.team).distinct().all()
    return [team[0] for team in teams]


@router.get("/assets", response_model=List[AssetResponse])
async def get_assets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    assets = db.query(Asset).order_by(Asset.organization, Asset.ip_address).all()
    return [AssetResponse.from_orm(a) for a in assets]


@router.get("/assets/organizations")
async def get_asset_organizations(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    orgs = db.query(Asset.organization).distinct().order_by(Asset.organization).all()
    return [o[0] for o in orgs]


@router.get("/assets/search", response_model=List[AssetResponse])
async def search_assets(
    keyword: Optional[str] = "",
    organization: Optional[str] = "",
    ip_range: Optional[str] = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Asset)
    
    if organization:
        query = query.filter(Asset.organization.like(f"%{organization}%"))
    
    if ip_range:
        try:
            import ipaddress
            network = ipaddress.ip_network(ip_range, strict=False)
            assets = []
            for asset in query.all():
                try:
                    ip = ipaddress.ip_address(asset.ip_address)
                    if ip in network:
                        assets.append(asset)
                except ValueError:
                    pass
            return [AssetResponse.from_orm(a) for a in assets]
        except ValueError:
            pass
    
    if keyword:
        query = query.filter(
            or_(
                Asset.ip_address.like(f"%{keyword}%"),
                Asset.organization.like(f"%{keyword}%"),
                Asset.purpose.like(f"%{keyword}%")
            )
        )
    
    return [AssetResponse.from_orm(a) for a in query.order_by(Asset.organization, Asset.ip_address).all()]


@router.post("/assets", response_model=AssetResponse)
async def create_asset(
    data: AssetCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    asset = Asset(
        ip_address=data.ip_address,
        organization=data.organization,
        purpose=data.purpose
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return AssetResponse.from_orm(asset)


@router.put("/assets/{asset_id}", response_model=AssetResponse)
async def update_asset(
    asset_id: int,
    data: AssetUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")
    
    if data.ip_address is not None:
        asset.ip_address = data.ip_address
    if data.organization is not None:
        asset.organization = data.organization
    if data.purpose is not None:
        asset.purpose = data.purpose
    
    db.commit()
    db.refresh(asset)
    return AssetResponse.from_orm(asset)


@router.delete("/assets/{asset_id}")
async def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")
    
    db.delete(asset)
    db.commit()
    
    return {"status": "ok", "message": "资产删除成功"}


@router.post("/assets/import")
async def import_assets(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import json
    import io
    
    content = await file.read()
    filename = file.filename.lower()
    
    assets_data = []
    
    if filename.endswith('.json'):
        try:
            data = json.loads(content.decode('utf-8'))
            if isinstance(data, list):
                assets_data = data
            elif isinstance(data, dict) and 'assets' in data:
                assets_data = data['assets']
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"JSON解析错误: {str(e)}")
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            ip_idx = headers.index('IP地址') if 'IP地址' in headers else None
            org_idx = headers.index('所属单位') if '所属单位' in headers else None
            purpose_idx = headers.index('主要用途') if '主要用途' in headers else None
            
            if ip_idx is None or org_idx is None:
                raise HTTPException(status_code=400, detail="Excel文件必须包含'IP地址'和'所属单位'列")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ip_idx] and row[org_idx]:
                    assets_data.append({
                        'ip_address': str(row[ip_idx]),
                        'organization': str(row[org_idx]),
                        'purpose': str(row[purpose_idx]) if purpose_idx and row[purpose_idx] else None
                    })
        except ImportError:
            raise HTTPException(status_code=500, detail="未安装openpyxl，请先安装：pip install openpyxl")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel解析错误: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="仅支持JSON和Excel文件")
    
    count = 0
    for item in assets_data:
        if 'ip_address' in item and 'organization' in item:
            asset = Asset(
                ip_address=item['ip_address'],
                organization=item['organization'],
                purpose=item.get('purpose')
            )
            db.add(asset)
            count += 1
    
    db.commit()
    return {"status": "ok", "message": f"成功导入 {count} 条资产数据"}


@router.get("/chat/users", response_model=List[dict])
async def get_chat_users(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    users = db.query(User).all()
    return [{"id": u.id, "username": u.username} for u in users]


@router.get("/chat/rooms", response_model=List[ChatRoomResponse])
async def get_chat_rooms(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rooms = db.query(ChatRoom).join(
        ChatRoomMember, ChatRoomMember.room_id == ChatRoom.id
    ).filter(
        ChatRoomMember.user_id == current_user.id
    ).order_by(ChatRoom.updated_at.desc()).all()
    result = []
    for room in rooms:
        members = []
        for member in room.members:
            members.append({
                "user_id": member.user_id,
                "username": member.user.username if member.user else "unknown",
                "joined_at": member.joined_at
            })
        result.append(ChatRoomResponse(
            id=room.id,
            name=room.name,
            description=room.description,
            created_by=room.created_by,
            created_by_username=room.created_by_user.username if room.created_by_user else None,
            created_at=room.created_at,
            updated_at=room.updated_at,
            unread_count=0,
            member_count=len(members),
            members=members
        ))
    return result


@router.post("/chat/rooms", response_model=ChatRoomResponse)
async def create_chat_room(request: ChatRoomCreateRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    room = ChatRoom(
        name=request.name,
        description=request.description,
        created_by=current_user.id
    )
    db.add(room)
    db.commit()
    db.refresh(room)

    member_ids = set(request.member_ids) if request.member_ids else set()
    member_ids.add(current_user.id)

    for user_id in member_ids:
        existing = db.query(ChatRoomMember).filter(
            ChatRoomMember.room_id == room.id,
            ChatRoomMember.user_id == user_id
        ).first()
        if not existing:
            member = ChatRoomMember(room_id=room.id, user_id=user_id)
            db.add(member)
    
    db.commit()

    members = []
    for member in room.members:
        members.append({
            "user_id": member.user_id,
            "username": member.user.username if member.user else "unknown",
            "joined_at": member.joined_at
        })

    return ChatRoomResponse(
        id=room.id,
        name=room.name,
        description=room.description,
        created_by=room.created_by,
        created_by_username=current_user.username,
        created_at=room.created_at,
        updated_at=room.updated_at,
        unread_count=0,
        member_count=len(members),
        members=members
    )


@router.get("/chat/rooms/{room_id}/messages", response_model=List[ChatMessageResponse])
async def get_chat_messages(room_id: int, limit: int = 50, offset: int = 0, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    is_member = db.query(ChatRoomMember).filter(
        ChatRoomMember.room_id == room_id,
        ChatRoomMember.user_id == current_user.id
    ).first()
    if not is_member:
        raise HTTPException(status_code=403, detail="无权访问该聊天室")
    
    messages = db.query(ChatMessage).filter(ChatMessage.room_id == room_id).order_by(ChatMessage.created_at.asc()).offset(offset).limit(limit).all()
    result = []
    for msg in messages:
        reply_content = None
        reply_username = None
        if msg.reply_to:
            reply_msg = db.query(ChatMessage).filter(ChatMessage.id == msg.reply_to).first()
            if reply_msg:
                reply_content = reply_msg.content[:50] + "..." if len(reply_msg.content) > 50 else reply_msg.content if reply_msg.content else ""
                if reply_msg.user:
                    reply_username = reply_msg.user.username
        result.append(ChatMessageResponse(
            id=msg.id,
            room_id=msg.room_id,
            user_id=msg.user_id,
            username=msg.user.username if msg.user else "unknown",
            content=msg.content,
            message_type=msg.message_type,
            file_path=msg.file_path,
            file_name=msg.file_name,
            file_size=msg.file_size,
            reply_to=msg.reply_to,
            reply_content=reply_content,
            reply_username=reply_username,
            created_at=msg.created_at
        ))
    return result


@router.post("/chat/messages", response_model=ChatMessageResponse)
async def send_chat_message(request: ChatMessageCreateRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    is_member = db.query(ChatRoomMember).filter(
        ChatRoomMember.room_id == request.room_id,
        ChatRoomMember.user_id == current_user.id
    ).first()
    if not is_member:
        raise HTTPException(status_code=403, detail="无权在该聊天室发送消息")
    
    msg = ChatMessage(
        room_id=request.room_id,
        user_id=current_user.id,
        content=request.content,
        message_type=request.message_type,
        file_path=request.file_path,
        file_name=request.file_name,
        file_size=request.file_size,
        reply_to=request.reply_to
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    
    room = db.query(ChatRoom).filter(ChatRoom.id == request.room_id).first()
    if room:
        room.updated_at = func.now()
        db.commit()
    
    reply_content = None
    reply_username = None
    if msg.reply_to:
        reply_msg = db.query(ChatMessage).filter(ChatMessage.id == msg.reply_to).first()
        if reply_msg:
            reply_content = reply_msg.content[:50] + "..." if len(reply_msg.content) > 50 else reply_msg.content if reply_msg.content else ""
            if reply_msg.user:
                reply_username = reply_msg.user.username
    
    return ChatMessageResponse(
        id=msg.id,
        room_id=msg.room_id,
        user_id=msg.user_id,
        username=current_user.username,
        content=msg.content,
        message_type=msg.message_type,
        file_path=msg.file_path,
        file_name=msg.file_name,
        file_size=msg.file_size,
        reply_to=msg.reply_to,
        reply_content=reply_content,
        reply_username=reply_username,
        created_at=msg.created_at
    )


@router.post("/chat/upload")
async def upload_chat_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    import os
    upload_dir = os.path.join(os.path.dirname(__file__), "..", "uploads", "chat")
    os.makedirs(upload_dir, exist_ok=True)
    
    import uuid
    file_ext = os.path.splitext(file.filename)[1]
    file_id = str(uuid.uuid4()) + file_ext
    file_path = os.path.join(upload_dir, file_id)
    
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    return {
        "file_path": f"/api/chat/download/{file_id}",
        "file_name": file.filename,
        "file_size": len(content)
    }


@router.get("/chat/download/{file_id}")
async def download_chat_file(file_id: str, current_user: User = Depends(get_current_user)):
    import os
    file_path = os.path.join(os.path.dirname(__file__), "..", "uploads", "chat", file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(file_path)


@router.post("/chat/rooms/{room_id}/members")
async def add_chat_member(room_id: int, request: AddChatMemberRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="聊天室不存在")
    
    if room.created_by != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="只有群主或管理员可以添加成员")
    
    existing = db.query(ChatRoomMember).filter(
        ChatRoomMember.room_id == room_id,
        ChatRoomMember.user_id == request.user_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户已在群中")
    
    target_user = db.query(User).filter(User.id == request.user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    member = ChatRoomMember(room_id=room_id, user_id=request.user_id)
    db.add(member)
    db.commit()
    
    return {"message": "添加成功"}


@router.delete("/chat/rooms/{room_id}/members/{user_id}")
async def remove_chat_member(room_id: int, user_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="聊天室不存在")
    
    if room.created_by != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="只有群主或管理员可以移除成员")
    
    if room.created_by == user_id:
        raise HTTPException(status_code=400, detail="不能移除群主")
    
    member = db.query(ChatRoomMember).filter(
        ChatRoomMember.room_id == room_id,
        ChatRoomMember.user_id == user_id
    ).first()
    if not member:
        raise HTTPException(status_code=404, detail="用户不在群中")
    
    db.delete(member)
    db.commit()
    
    return {"message": "移除成功"}


@router.delete("/chat/rooms/{room_id}")
async def delete_chat_room(room_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    room = db.query(ChatRoom).filter(ChatRoom.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="聊天室不存在")
    
    if room.created_by != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="只有群主或管理员可以删除群")
    
    db.query(ChatMessage).filter(ChatMessage.room_id == room_id).delete()
    db.query(ChatRoomMember).filter(ChatRoomMember.room_id == room_id).delete()
    db.delete(room)
    db.commit()
    
    return {"message": "删除成功"}
